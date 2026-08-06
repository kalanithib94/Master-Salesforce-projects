# -*- coding: utf-8 -*-
"""
Export E2E skill results + field help guide for Google Sheets / Excel (CSV).

Reads:
  - Reports/MAIN_REPORT.html (latest outcomes)
  - Deliverables/docs/PROMPT_COMMANDS_BY_SKILL.json (fields / help text)
Writes:
  - Reports/exports/*.csv
  - Optionally pushed to Google Sheet via write_to_google_sheet() helper data files
"""
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_main_report import next_step_for_skill, parse_seeded_html  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
LIBRARY = ROOT.parent
REPORTS = LIBRARY / "Reports"
EXPORTS = REPORTS / "exports"
MAIN = REPORTS / "MAIN_REPORT.html"
PROMPT_JSON = LIBRARY / "Deliverables" / "docs" / "PROMPT_COMMANDS_BY_SKILL.md"
PROMPT_JSON_ALT = LIBRARY / "Deliverables" / "docs" / "PROMPT_COMMANDS_BY_SKILL.json"
SEEDED = Path(__file__).resolve().parent / "results" / "matrix_report_seeded.json"

# Sheet ids — also documented in Reports/exports/README_SHEET.md
SHEET_ID = "1LHw46KlgmFam2cX5sMBPSIxEYNrEUqx7E22ihp-9K5E"  # index
FULL_SHEET_ID = "1LsHePYKTJ5rn3OYQpyud3lX02MJ2SijRH3cZcCpo9vI"  # full values

# Infer primary Salesforce object from skill name (help for users)
OBJECT_HINT = {
    "account": "Account",
    "contact": "Contact",
    "lead": "Lead",
    "opportunit": "Opportunity",
    "case": "Case",
    "campaign": "Campaign",
    "product": "Product2",
    "pricebook": "PricebookEntry",
    "quote": "Quote",
    "order": "Order",
    "contract": "Contract",
    "asset": "Asset",
    "task": "Task",
    "event": "Event",
    "work_order": "WorkOrder",
    "service_appointment": "ServiceAppointment",
    "partner": "Account (Partner)",
    "queue": "Group (Queue) / Case",
    "knowledge": "KnowledgeArticleVersion",
    "subscription": "Subscription (feature)",
    "cpq": "SBQQ__Quote__c (CPQ)",
    "care": "CarePlan / CareTask",
    "financial": "Financial Account",
    "session": "Session context",
    "picklist": "Describe (any object)",
    "transfer": "Any record (OwnerId)",
    "assign_to_queue": "Case / Lead",
    "activity": "Task",
    "approvals": "ProcessInstance",
}


def clip(s: object, n: int = 4500) -> str:
    t = "" if s is None else str(s)
    t = re.sub(r"\s+", " ", t).strip()
    return t if len(t) <= n else t[: n - 3] + "..."


def primary_object(skill: str) -> str:
    sk = skill.lower()
    for k, v in OBJECT_HINT.items():
        if k in sk:
            return v
    return "Multiple / generic"


def what_happened(category: str, response: str) -> str:
    if category == "pass":
        return "Skill executed successfully (Apex business success)."
    if category == "fail_missing_feature":
        return "Feature / package not available in this org (expected N/A until enabled)."
    if category == "fail_data":
        return "Data / seed or missing parameter issue: " + clip(response, 400)
    if category == "fail_business":
        return "Handler returned a business error: " + clip(response, 400)
    if category == "fail_api":
        return "API / transport failure: " + clip(response, 400)
    if category == "fail_missing_class":
        return "Skill or Apex class not linked / not deployed."
    return clip(response, 400)


def load_prompt_map() -> dict:
    path = PROMPT_JSON_ALT if PROMPT_JSON_ALT.exists() else None
    if path is None:
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def format_fields(skill: str, cmd: dict | None) -> tuple[str, str, str, str]:
    """Returns (required, fields_help, related_hints, user_expect)."""
    if not cmd:
        return "", "See Prompt Command in org / package seed.", "", "Use skill via agent or invokeAgentSkill with mapped params."
    req = cmd.get("required") or []
    props = cmd.get("properties") or []
    lines = []
    related = set()
    if isinstance(props, list):
        items = props
    elif isinstance(props, dict):
        items = [{"name": k, **(v if isinstance(v, dict) else {"description": str(v)})} for k, v in props.items()]
    else:
        items = []
    for p in items:
        name = p.get("name") or ""
        typ = p.get("type") or "string"
        desc = (p.get("description") or "").strip()
        flag = "REQUIRED" if name in req else "optional"
        lines.append(f"{name} ({typ}, {flag}): {desc}")
        low = f"{name} {desc}".lower()
        for token, obj in (
            ("account", "Account"),
            ("contact", "Contact"),
            ("lead", "Lead"),
            ("opportunit", "Opportunity"),
            ("case", "Case"),
            ("campaign", "Campaign"),
            ("pricebook", "Pricebook2 / PricebookEntry"),
            ("product", "Product2"),
            ("quote", "Quote"),
            ("order", "Order"),
            ("user", "User"),
            ("queue", "Queue (Group)"),
        ):
            if token in low:
                related.add(obj)
    fields_help = " | ".join(lines) if lines else "(schema empty / additionalProperties)"
    req_s = ", ".join(req) if req else "(none listed)"
    rel_s = ", ".join(sorted(related)) if related else primary_object(skill)
    expect = (
        f"As a user: ask the agent in plain language (or pass JSON params). "
        f"Primary object: {primary_object(skill)}. "
        f"Prefer names / CaseNumber when you do not know Salesforce Ids. "
        f"Mutating skills usually need a confirmation step first."
    )
    return req_s, clip(fields_help, 4500), rel_s, expect


def load_results() -> list[dict]:
    rows = parse_seeded_html(MAIN) if MAIN.exists() else []
    # enrich from seeded json when available
    by_skill: dict[str, dict] = {}
    if SEEDED.exists():
        data = json.loads(SEEDED.read_text(encoding="utf-8"))
        for r in data.get("results") or []:
            by_skill[r.get("skill") or ""] = r
    out = []
    for r in rows:
        skill = r.get("skill") or ""
        rich = by_skill.get(skill, {})
        req = r.get("request")
        if not isinstance(req, str):
            req = json.dumps(req or rich.get("request") or {}, indent=2, ensure_ascii=False)
        elif rich.get("request") and req.strip() in ("{}", ""):
            req = json.dumps(rich.get("request"), indent=2, ensure_ascii=False)
        resp = r.get("response") or ""
        if not resp and rich.get("errorSnippet"):
            resp = str(rich.get("errorSnippet"))
        if not resp and rich.get("apexData") is not None:
            ad = rich["apexData"]
            if isinstance(ad, dict):
                resp = ad.get("message") or ad.get("error") or json.dumps(ad, ensure_ascii=False)
            else:
                resp = str(ad)
        cat = r.get("category") or rich.get("category") or "unknown"
        out.append(
            {
                "skill": skill,
                "category": cat,
                "request": req,
                "response": resp,
                "http": rich.get("http"),
                "elapsedSec": rich.get("elapsedSec"),
            }
        )
    if out:
        return sorted(out, key=lambda x: x["skill"])
    # fallback seeded only
    for skill, rich in sorted(by_skill.items()):
        req = json.dumps(rich.get("request") or {}, indent=2, ensure_ascii=False)
        ad = rich.get("apexData")
        if isinstance(ad, dict):
            resp = ad.get("message") or ad.get("error") or json.dumps(ad, ensure_ascii=False)
        else:
            resp = str(ad or rich.get("errorSnippet") or "")
        out.append(
            {
                "skill": skill,
                "category": rich.get("category") or "unknown",
                "request": req,
                "response": resp,
                "http": rich.get("http"),
                "elapsedSec": rich.get("elapsedSec"),
            }
        )
    return out


def build_tables() -> dict[str, list[list[str]]]:
    results = load_results()
    prompts = load_prompt_map()
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    counts = Counter(r["category"] for r in results)

    summary = [
        ["GPTfy Agent Skills — E2E + Field Guide", ""],
        ["Updated", now],
        ["Source", "MAIN_REPORT.html + PROMPT_COMMANDS_BY_SKILL.json"],
        ["Org / Note", "Master Dev multi-skill matrix (strict Apex business success)"],
        ["Total skills", str(len(results))],
        ["pass", str(counts.get("pass", 0))],
        ["fail_data", str(counts.get("fail_data", 0))],
        ["fail_business", str(counts.get("fail_business", 0))],
        ["fail_missing_feature", str(counts.get("fail_missing_feature", 0))],
        ["fail_api", str(counts.get("fail_api", 0))],
        ["fail_missing_class", str(counts.get("fail_missing_class", 0))],
        ["", ""],
        ["How to read", ""],
        ["E2E Results tab", "One row per skill: last request payload, response, result, next step"],
        ["Skill Field Guide tab", "User-facing parameters, objects, and what to expect (help text)"],
        ["HTML report", "AGENT LIBRARY/Reports/MAIN_REPORT.html (full UI detail)"],
    ]

    e2e_header = [
        "Skill",
        "Status",
        "Primary object",
        "Request (JSON)",
        "Response (summary)",
        "What happened",
        "Next step / what needs to be done",
        "HTTP",
        "Elapsed (s)",
    ]
    e2e = [e2e_header]
    for r in results:
        cat = r["category"]
        resp = clip(r["response"], 3500)
        e2e.append(
            [
                r["skill"],
                cat,
                primary_object(r["skill"]),
                clip(r["request"], 3500),
                resp,
                what_happened(cat, resp),
                next_step_for_skill(r["skill"], cat, resp) or ("—" if cat == "pass" else "Review response and re-test"),
                "" if r.get("http") is None else str(r.get("http")),
                "" if r.get("elapsedSec") is None else str(r.get("elapsedSec")),
            ]
        )

    guide_header = [
        "Skill",
        "Primary object",
        "Related / used objects",
        "Required params",
        "Field / parameter help (what each field means)",
        "What the user can expect (help text)",
    ]
    guide = [guide_header]
    # Union of matrix skills + all prompt commands
    skills = sorted(set([r["skill"] for r in results]) | set(prompts.keys()))
    for skill in skills:
        cmd = prompts.get(skill)
        req_s, fields_help, rel_s, expect = format_fields(skill, cmd)
        guide.append(
            [
                skill,
                primary_object(skill),
                rel_s,
                req_s,
                fields_help,
                expect,
            ]
        )

    return {"Summary": summary, "E2E Results": e2e, "Skill Field Guide": guide}


def write_csv(tables: dict[str, list[list[str]]]) -> list[Path]:
    import csv

    EXPORTS.mkdir(parents=True, exist_ok=True)
    paths = []
    for name, rows in tables.items():
        p = EXPORTS / f"{name.replace(' ', '_').lower()}.csv"
        with p.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            for row in rows:
                w.writerow(row)
        paths.append(p)
        print("CSV:", p, "rows", len(rows))
    # bundle payload for MCP / external push
    payload = {k: v for k, v in tables.items()}
    jp = EXPORTS / "sheet_payload.json"
    jp.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print("JSON payload:", jp)
    return paths


def write_xlsx(tables: dict[str, list[list[str]]]) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed; skip xlsx (pip install openpyxl)")
        return None

    wb = Workbook()
    first = True
    for title, rows in tables.items():
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        for r in rows:
            ws.append([("" if c is None else str(c)[:32000]) for c in r])
        for col in range(1, min(12, (ws.max_column or 1) + 1)):
            ws.column_dimensions[get_column_letter(col)].width = 22
    out = EXPORTS / "GPTfy_Skill_Library_E2E.xlsx"
    EXPORTS.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print("XLSX:", out)
    return out


def main() -> int:
    tables = build_tables()
    write_csv(tables)
    write_xlsx(tables)
    print("Sheets: see Reports/exports/README_SHEET.md")
    print("FULL", "https://docs.google.com/spreadsheets/d/{}/edit".format(FULL_SHEET_ID))
    print("INDEX", "https://docs.google.com/spreadsheets/d/{}/edit".format(SHEET_ID))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
